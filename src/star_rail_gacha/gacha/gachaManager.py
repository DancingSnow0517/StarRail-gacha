import json
import os.path
from typing import List, Dict, Tuple

from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .gacha import Gacha
from .types import GachaType


class GachaManager:

    def __init__(self, uid: str):
        self.uid = uid
        self.records: Dict[GachaType, List[Gacha]] = {}

    def get_gacha_list(self, gacha_type: GachaType) -> List[Gacha]:
        return self.records.get(gacha_type)

    def save_to_file(self, path=None):
        if not os.path.exists('userData'):
            os.mkdir('userData')
        path = f'userData/gacha-list-{self.uid}.json' if path is None else path
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(self.record_dict, f, indent=4, ensure_ascii=False)

    def load_from_file(self):
        if not os.path.exists('userData'):
            os.mkdir('userData')
        if os.path.exists(f'userData/gacha-list-{self.uid}.json'):
            with open(f'userData/gacha-list-{self.uid}.json', 'r', encoding='utf-8') as f:
                data = json.load(f)
            for gacha_type in GachaType:
                self.records[gacha_type] = []
                for d in data.get(gacha_type.value, []):
                    self.records[gacha_type].append(Gacha(**d))
        else:
            for gacha_type in GachaType:
                self.records[gacha_type] = []
            self.save_to_file()

    def add_records(self, records: List[Gacha]) -> Tuple[bool, int]:
        flag = True
        count = 0
        for record in records:
            if self.add_record(record):
                count += 1
            else:
                flag = False
        return flag, count

    def add_record(self, record: Gacha) -> bool:
        if self.is_record_exist(record):
            return False
        self.records[record.gacha_type].append(record)
        sorted(self.records[record.gacha_type], key=lambda x: x.id, reverse=True)
        return True

    def is_record_exist(self, record: Gacha) -> bool:
        for r in self.records[record.gacha_type]:
            if r.id == record.id:
                return True
        return False

    def get_count(self, gacha_type: GachaType) -> int:
        return len(self.records[gacha_type])

    def get_5star_count(self, gacha_type: GachaType) -> int:
        count = 0
        for record in self.records[gacha_type]:
            if record.is_5star:
                count += 1
        return count

    def get_5star_character_count(self, gacha_type: GachaType) -> int:
        count = 0
        for record in self.records[gacha_type]:
            if record.is_5star and record.is_character:
                count += 1
        return count

    def get_5star_light_cone_count(self, gacha_type: GachaType) -> int:
        count = 0
        for record in self.records[gacha_type]:
            if record.is_5star and record.is_light_cone:
                count += 1
        return count

    def get_4star_count(self, gacha_type: GachaType) -> int:
        count = 0
        for record in self.records[gacha_type]:
            if record.is_4star:
                count += 1
        return count

    def get_4star_character_count(self, gacha_type: GachaType) -> int:
        count = 0
        for record in self.records[gacha_type]:
            if record.is_4star and record.is_character:
                count += 1
        return count

    def get_4star_light_cone_count(self, gacha_type: GachaType) -> int:
        count = 0
        for record in self.records[gacha_type]:
            if record.is_4star and record.is_light_cone:
                count += 1
        return count

    def get_3star_count(self, gacha_type: GachaType) -> int:
        count = 0
        for record in self.records[gacha_type]:
            if record.is_3star:
                count += 1
        return count

    def get_last_5star_count(self, gacha_type: GachaType) -> int:
        count = 0
        for record in self.records[gacha_type]:
            if record.is_5star:
                break
            count += 1
        return count

    def get_last_4star_count(self, gacha_type: GachaType) -> int:
        count = 0
        for record in self.records[gacha_type]:
            if record.is_4star:
                break
            count += 1
        return count

    def get_5star_history(self, gacha_type: GachaType) -> List[Tuple[str, int]]:
        rt = []
        cost = 0
        for record in reversed(self.records[gacha_type]):
            if not record.is_5star:
                cost += 1
            else:
                cost += 1
                rt.append((record.name, cost))
                cost = 0
        return rt

    def get_5star_average(self, gacha_type: GachaType) -> float:
        count = 0
        total = 0
        for _, cost in self.get_5star_history(gacha_type):
            count += 1
            total += cost
        if total == 0:
            return 0
        return total / count

    def save_to_excel(self, path: str):
        wb = Workbook()
        for gacha_type in GachaType:
            if gacha_type == GachaType.CHARACTER:
                sheet_name = '角色活动跃迁'
            elif gacha_type == GachaType.LIGHT_CONE:
                sheet_name = '光锥活动跃迁'
            elif gacha_type == GachaType.STELLAR:
                sheet_name = '群星跃迁'
            else:
                sheet_name = '始发跃迁'
            ws = wb.create_sheet(sheet_name)  # type: Worksheet
            ws.append(['时间', '名称', '类别', '星级', '总跃迁次数', '保底内'])
            for col in range(1, 7):
                ws.cell(row=1, column=col).fill = PatternFill(fill_type='solid', fgColor='DBD7D3')
                ws.cell(row=1, column=col).border = Border(left=Side(border_style='thin', color='D9D9D9'),
                                                           right=Side(border_style='thin', color='D9D9D9'),
                                                           top=Side(border_style='thin', color='D9D9D9'),
                                                           bottom=Side(border_style='thin', color='D9D9D9'))
                ws.cell(row=1, column=col).font = Font(bold=True, color='757595', name='微软雅黑')
            cost = 0
            for i, record in enumerate(reversed(self.get_gacha_list(gacha_type))):
                cost += 1
                ws.append([record.time, record.name, record.item_type.value, record.rank_type, i + 1, cost])
                row = i + 2
                for col in range(1, 7):
                    ws.cell(row=row, column=col).fill = PatternFill(fill_type='solid', fgColor='EBEBEB')
                    ws.cell(row=row, column=col).border = Border(left=Side(border_style='thin', color='D9D9D9'),
                                                                 right=Side(border_style='thin', color='D9D9D9'),
                                                                 top=Side(border_style='thin', color='D9D9D9'),
                                                                 bottom=Side(border_style='thin', color='D9D9D9'))
                    if record.is_5star:
                        ws.cell(row=row, column=col).font = Font(bold=True, color='BD6932', name='微软雅黑')
                        cost = 0
                    elif record.is_4star:
                        ws.cell(row=row, column=col).font = Font(bold=True, color='A256E1', name='微软雅黑')
                    else:
                        ws.cell(row=row, column=col).font = Font(color='8E8E8E', name='微软雅黑')
            ws.column_dimensions['A'].width = 24
            ws.column_dimensions['B'].width = 13
            ws.column_dimensions['C'].width = 7
            ws.column_dimensions['D'].width = 7
            ws.column_dimensions['E'].width = 11
            ws.column_dimensions['F'].width = 7
        wb.remove(wb['Sheet'])
        wb.save(path)

    def clear(self):
        for gacha_type in GachaType:
            self.records[gacha_type] = []
        self.save_to_file()

    @classmethod
    def load_from_uid(cls, uid: str):
        rt = cls(uid)
        rt.load_from_file()
        return rt

    @property
    def record_dict(self):
        rt = {}
        for gacha_type in GachaType:
            rt[gacha_type.value] = []
            for record in self.records[gacha_type]:
                rt[gacha_type.value].append(record.dict)
        return rt
